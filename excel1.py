import openpyxl
#wb = openpyxl.load_workbook("example.xlsx")
#print(type(wb))
#print(wb.get_sheet_names()) # to get the sheets names from the workbook
#sheet = wb.get_sheet_by_name('Sheet3')
#print(sheet)
#print(sheet.title) # getting title of the sheet object
#anotherSheet = wb.active # to get the active sheet
#print(anotherSheet)
#sheet1 = wb.get_sheet_by_name('Sheet1')
#print(sheet1["A1"])
#print(sheet1["A1"].value) # to get the value of the sheet
#c = sheet1["B1"]
#print(c.value)
#print("Row" +str(c.row) +", Column" + str(c.column) + " is " + str(c.value))
#print('Cell ' + c.coordinate + ' is ' + c.value)
#print(sheet1["C1"].value)
#a=sheet1.cell(row =2, column = 2)
#print(a.value)
#print(type(a))
#for i in range(1,8): # to get all the values of a cloumn
#    print(i,sheet1.cell(row =i, column = 2).value)

#print(sheet1.max_row ,"*", sheet1.max_column) # to get the max no of columns and rows

# getting rows and columns from the sheets
#print(tuple(sheet1["A1":"C3"]))#
#for i in sheet1['A1':'C3']: # printing the values of rows and columns
#    for j in i:
#        print(j.coordinate,j.value,end=' ')
#    print(" ")
#    print("end of row")
wb = openpyxl.load_workbook("first_excel_file.xlsx")
ws=wb.active # to get the active worksheet in workbook
print(ws)
#ws1 = wb.get_sheet_by_name("data")
ws2 = wb.create_sheet("Mysheet") # creating a worksheet of name Mysheet
ws3 = wb.create_sheet("Mysheet",0) # inserting the sheet at position 0
ws2.title = "New Title"
ws5 = wb.create_sheet("Mysheet",0)
ws2.sheet_properties.tabColor = "1072BA" # to give the color to the sheet
ws4 = wb["New Title"]
print(wb.sheetnames)
print(wb.path)
wb.save("first_excel_file.xlsx")

