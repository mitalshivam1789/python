#reading data from a spread sheet

# readCensusExcel.py - Tabulates population and number of census tracts for each country

import openpyxl, pprint
print("Opening workbook...")
wb = openpyxl.load_workbook('censuspopdata.xlsx')
sheet = wb.get_sheet_by_name('Population by Census Tract')
countyData = {}
print('Reading rows.....')
print(sheet.max_row,sheet.max_column)

# -------------to print the rows --------------------#
st=[]
co=[]
po=[]

for row in range(2,sheet.max_row +1):
    state = sheet['B' + str(row)].value
    county = sheet['C' + str(row)].value
    pop = sheet['D' + str(row)].value
    st.append(state)
    co.append(county)
    po.append(pop)

#---------no of countries
st = set(st)
#---------


print(sheet.max_row,sheet.max_column)

 # --- Populate the data structure ----------------#





