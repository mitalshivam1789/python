import os
import openpyxl
print(openpyxl.__version__)
os.chdir('C:\\users\\HP\\Desktop')
wb=openpyxl.load_workbook('new.xlsx')# to load the workbook
print(type(wb))
print(wb)
print(wb.get_sheet_names())# to get the names of all the sheets in a workbook
sheet = wb.get_sheet_by_name('My Sheet')# to get the sheet by its name
print(type(sheet)) # to print the type of the sheet
print(sheet['A1'].value)# to get the value of the cell of the sheet
print(sheet['A3'].value)
sheet['C3'].value = 42 # changing the value of the cell of the sheet
wb.save('example.xlsx') # saving the data of the sheet into the different workbook naming it example.xlsx
print(sheet.title) # to get the title of the sheet
#sheet.title="My Sheet" # to change the title of the sheet
print(sheet.title)
wb.save("new.xlsx")
#----way to print all the values of the columns
print(sheet.cell(row=1, column=3).value)
for i in range(1,5): print(sheet.cell(row = i, column =3).value)
# --------------------
#print(openpyxl.cell.column_index_from_string('AA'))
#wb.create_sheet(title='New Sheet',index = 2)
#wb.save('example2.xlsx')

wb=openpyxl.load_workbook('example2.xlsx')
sheet = wb.get_sheet_by_name('My Sheet')
sheet.row_dimensions[1].height = 70# to change the height
sheet.column_dimensions['B'].width = 20 # to change the width
wb.save('example2.xlsx')
#sheet['C8'].value='=SUM(C1:C7)' # it will print the sum from column c1 to c7 and print it in c8 cell


